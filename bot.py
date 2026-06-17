import os
import re
import json
import asyncio
import shutil
import zipfile
from datetime import datetime, timedelta, time
import pytz
from dotenv import load_dotenv
from telegram import Update, BotCommand, BotCommandScopeChat
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import asyncpg
import tempfile

load_dotenv()

TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID"))
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "silverzen")
DATABASE_URL = os.getenv("DATABASE_URL")

# Время МСК+1 (UTC+4)
MSK_PLUS_1 = pytz.timezone('Europe/Samara')

# Глобальная переменная для пула подключений к БД
db_pool = None

# ==================== РАБОТА С БАЗОЙ ДАННЫХ ====================
async def init_db():
    """Инициализация подключения к БД и создание таблиц"""
    global db_pool
    
    if db_pool is None:
        db_pool = await asyncpg.create_pool(DATABASE_URL, min_size=1, max_size=5)
        print("✅ Подключение к PostgreSQL установлено")
    
    # Создаём таблицы
    async with db_pool.acquire() as conn:
        # Таблица пользователей
        await conn.execute('''
            CREATE TABLE IF NOT EXISTS users (
                user_id BIGINT PRIMARY KEY,
                username TEXT,
                joined TIMESTAMP,
                status TEXT DEFAULT 'active',
                days_count INT DEFAULT 1,
                last_reminder_sent DATE,
                access_until DATE,
                payment_info TEXT,
                payment_date DATE
            )
        ''')
        
        # Таблица давления
        await conn.execute('''
            CREATE TABLE IF NOT EXISTS pressure (
                id SERIAL PRIMARY KEY,
                user_id BIGINT REFERENCES users(user_id) ON DELETE CASCADE,
                date DATE NOT NULL,
                time TIME NOT NULL,
                period TEXT CHECK (period IN ('Утро', 'День', 'Вечер')),
                systolic INT,
                diastolic INT,
                pulse INT,
                comment TEXT,
                created_at TIMESTAMP DEFAULT NOW()
            )
        ''')
        
        # Таблица глюкозы
        await conn.execute('''
            CREATE TABLE IF NOT EXISTS glucose (
                id SERIAL PRIMARY KEY,
                user_id BIGINT REFERENCES users(user_id) ON DELETE CASCADE,
                date DATE NOT NULL,
                time TIME NOT NULL,
                period TEXT CHECK (period IN ('Утро', 'День', 'Вечер')),
                glucose_value DECIMAL(4,1),
                glucose_type TEXT,
                comment TEXT,
                created_at TIMESTAMP DEFAULT NOW()
            )
        ''')
        
        # Индексы для ускорения запросов
        await conn.execute('CREATE INDEX IF NOT EXISTS idx_pressure_user_id ON pressure(user_id)')
        await conn.execute('CREATE INDEX IF NOT EXISTS idx_glucose_user_id ON glucose(user_id)')
        
        print("✅ Таблицы созданы/проверены")

async def get_db_pool():
    """Возвращает пул подключений к БД"""
    return db_pool

# ==================== РАБОТА С ПОЛЬЗОВАТЕЛЯМИ ====================
async def add_user(user_id, username):
    users = await get_all_users()
    if str(user_id) not in users:
        now = datetime.now(MSK_PLUS_1)
        async with db_pool.acquire() as conn:
            await conn.execute('''
                INSERT INTO users (user_id, username, joined, status, days_count)
                VALUES ($1, $2, $3, 'active', 1)
            ''', user_id, username, now)
        return True
    else:
        # Обновляем username если изменился
        if users.get(str(user_id), {}).get("username") != username:
            async with db_pool.acquire() as conn:
                await conn.execute('UPDATE users SET username = $1 WHERE user_id = $2', username, user_id)
    return False

async def get_all_users():
    """Возвращает всех пользователей в виде словаря (как раньше в users.json)"""
    async with db_pool.acquire() as conn:
        rows = await conn.fetch('SELECT * FROM users')
        users = {}
        for row in rows:
            uid = str(row['user_id'])
            users[uid] = {
                "username": row['username'],
                "joined": row['joined'].strftime("%d-%m-%Y %H:%M:%S") if row['joined'] else None,
                "status": row['status'],
                "days_count": row['days_count'],
                "last_reminder_sent": row['last_reminder_sent'].strftime("%d-%m-%Y") if row['last_reminder_sent'] else "",
                "access_until": row['access_until'].strftime("%d-%m-%Y") if row['access_until'] else None,
                "payment_info": row['payment_info'],
                "payment_date": row['payment_date'].strftime("%d-%m-%Y") if row['payment_date'] else None
            }
        return users

async def update_user_days(user_id):
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow('SELECT joined FROM users WHERE user_id = $1', user_id)
        if row:
            joined = row['joined']
            days = (datetime.now(MSK_PLUS_1) - joined).days + 1
            await conn.execute('UPDATE users SET days_count = $1 WHERE user_id = $2', days, user_id)
            return days
    return 0

async def check_and_send_3day_reminder(user_id, app):
    # Админа пропускаем
    if str(user_id) == str(ADMIN_ID):
        return False
    
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow('SELECT joined, status, last_reminder_sent FROM users WHERE user_id = $1', user_id)
        if row and row['status'] == 'active':
            joined = row['joined']
            days = (datetime.now(MSK_PLUS_1) - joined).days + 1
            last_reminder = row['last_reminder_sent']
            
            if days >= 3 and last_reminder != datetime.now(MSK_PLUS_1).date():
                await conn.execute('UPDATE users SET last_reminder_sent = $1 WHERE user_id = $2', 
                                  datetime.now(MSK_PLUS_1).date(), user_id)
                try:
                    await app.bot.send_message(
                        chat_id=int(user_id),
                        text="Вы уже 3 дня пользуетесь ботом. Если хотите продолжить, есть предложения или замечания, свяжитесь с админом."
                    )
                except:
                    pass
                return True
    return False

async def check_access(user_id):
    if str(user_id) == str(ADMIN_ID):
        return True
    
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow('SELECT status, access_until FROM users WHERE user_id = $1', user_id)
        if not row:
            return False
        
        if row['status'] == 'active':
            return True
        
        if row['status'] == 'access':
            if row['access_until']:
                if datetime.now(MSK_PLUS_1).date() <= row['access_until']:
                    return True
                else:
                    await conn.execute('UPDATE users SET status = $1 WHERE user_id = $2', 'blocked', user_id)
                    return False
            return True
    
    return False

async def grant_access(user_id, days):
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow('SELECT * FROM users WHERE user_id = $1', user_id)
        if row:
            access_until = (datetime.now(MSK_PLUS_1) + timedelta(days=days)).date()
            await conn.execute('UPDATE users SET status = $1, access_until = $2 WHERE user_id = $3', 
                              'access', access_until, user_id)
            return True
    return False

def is_admin(user_id):
    return user_id == ADMIN_ID

# ==================== ОПРЕДЕЛЕНИЕ ПЕРИОДА ПО ВРЕМЕНИ ====================
def get_period_by_time():
    now = datetime.now(MSK_PLUS_1)
    hour = now.hour
    
    if 6 <= hour < 12:
        return "Утро"
    elif 12 <= hour < 18:
        return "День"
    else:
        return "Вечер"

# ==================== РАБОТА С EXCEL ====================
async def generate_pressure_excel(user_id):
    """Генерирует Excel файл с давлением для конкретного пользователя"""
    wb = Workbook()
    
    # Удаляем дефолтный лист
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Лист 1: Давление
    ws_pressure = wb.create_sheet("Давление")
    headers_pressure = ['Дата', 'Время', 'Период', 'Верхнее', 'Нижнее', 'Пульс', 'Комментарий']
    for col, header in enumerate(headers_pressure, 1):
        cell = ws_pressure.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    col_widths_pressure = {'A': 12, 'B': 10, 'C': 8, 'D': 10, 'E': 10, 'F': 8, 'G': 40}
    for col_letter, width in col_widths_pressure.items():
        ws_pressure.column_dimensions[col_letter].width = width
    ws_pressure.row_dimensions[1].height = 20
    
    # Лист 2: Глюкоза
    ws_glucose = wb.create_sheet("Глюкоза")
    headers_glucose = ['Дата', 'Время', 'Период', 'Глюкоза', 'Тип замера', 'Комментарий']
    for col, header in enumerate(headers_glucose, 1):
        cell = ws_glucose.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    col_widths_glucose = {'A': 12, 'B': 10, 'C': 8, 'D': 10, 'E': 25, 'F': 40}
    for col_letter, width in col_widths_glucose.items():
        ws_glucose.column_dimensions[col_letter].width = width
    ws_glucose.row_dimensions[1].height = 20
    
    # Заполняем данными из БД
    async with db_pool.acquire() as conn:
        # Давление
        rows = await conn.fetch('''
            SELECT date, time, period, systolic, diastolic, pulse, comment 
            FROM pressure 
            WHERE user_id = $1 
            ORDER BY date DESC, time DESC
        ''', user_id)
        
        row_num = 2
        for row in rows:
            ws_pressure.cell(row=row_num, column=1, value=row['date'].strftime("%d-%m-%Y"))
            ws_pressure.cell(row=row_num, column=2, value=row['time'].strftime("%H:%M:%S"))
            ws_pressure.cell(row=row_num, column=3, value=row['period'])
            ws_pressure.cell(row=row_num, column=4, value=row['systolic'])
            ws_pressure.cell(row=row_num, column=5, value=row['diastolic'])
            ws_pressure.cell(row=row_num, column=6, value=row['pulse'] if row['pulse'] else "")
            ws_pressure.cell(row=row_num, column=7, value=row['comment'] if row['comment'] else "")
            row_num += 1
        
        # Глюкоза
        rows = await conn.fetch('''
            SELECT date, time, period, glucose_value, glucose_type, comment 
            FROM glucose 
            WHERE user_id = $1 
            ORDER BY date DESC, time DESC
        ''', user_id)
        
        row_num = 2
        for row in rows:
            ws_glucose.cell(row=row_num, column=1, value=row['date'].strftime("%d-%m-%Y"))
            ws_glucose.cell(row=row_num, column=2, value=row['time'].strftime("%H:%M:%S"))
            ws_glucose.cell(row=row_num, column=3, value=row['period'])
            ws_glucose.cell(row=row_num, column=4, value=float(row['glucose_value']))
            ws_glucose.cell(row=row_num, column=5, value=row['glucose_type'] if row['glucose_type'] else "")
            ws_glucose.cell(row=row_num, column=6, value=row['comment'] if row['comment'] else "")
            row_num += 1
    
    # Сохраняем во временный файл
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        wb.save(tmp.name)
        return tmp.name

# ==================== РАБОТА С ДАННЫМИ ====================
async def save_pressure_to_db(user_id, period, systolic, diastolic, pulse, comment):
    now = datetime.now(MSK_PLUS_1)
    
    if now.hour < 6:
        date_val = (now - timedelta(days=1)).date()
    else:
        date_val = now.date()
    
    time_val = now.time()
    
    async with db_pool.acquire() as conn:
        await conn.execute('''
            INSERT INTO pressure (user_id, date, time, period, systolic, diastolic, pulse, comment)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
        ''', user_id, date_val, time_val, period, systolic, diastolic, pulse, comment)

async def save_glucose_to_db(user_id, period, glucose, glucose_type, comment):
    now = datetime.now(MSK_PLUS_1)
    
    if now.hour < 6:
        date_val = (now - timedelta(days=1)).date()
    else:
        date_val = now.date()
    
    time_val = now.time()
    
    async with db_pool.acquire() as conn:
        await conn.execute('''
            INSERT INTO glucose (user_id, date, time, period, glucose_value, glucose_type, comment)
            VALUES ($1, $2, $3, $4, $5, $6, $7)
        ''', user_id, date_val, time_val, period, glucose, glucose_type, comment)

async def get_today_pressure_report(user_id):
    today = datetime.now(MSK_PLUS_1).date()
    
    async with db_pool.acquire() as conn:
        rows = await conn.fetch('''
            SELECT date, time, period, systolic, diastolic, pulse, comment 
            FROM pressure 
            WHERE user_id = $1 AND date = $2
            ORDER BY time ASC
        ''', user_id, today)
    
    if not rows:
        return f"📊 Отчет по давлению за {today.strftime('%d-%m-%Y')}\n\nНет данных."
    
    report = f"📊 Отчет по давлению за {today.strftime('%d-%m-%Y')}\n\n"
    for row in rows:
        period_emoji = {"Утро": "🌅", "День": "☀️", "Вечер": "🌙"}
        emoji = period_emoji.get(row['period'], "")
        
        report += f"{emoji} {row['period']} {row['time'].strftime('%H:%M:%S')}: {row['systolic']}/{row['diastolic']}"
        if row['pulse']:
            report += f", пульс {row['pulse']}"
        if row['comment']:
            report += f"\n   📝 {row['comment']}"
        report += "\n\n"
    
    return report

async def get_today_glucose_report(user_id):
    today = datetime.now(MSK_PLUS_1).date()
    
    async with db_pool.acquire() as conn:
        rows = await conn.fetch('''
            SELECT date, time, period, glucose_value, glucose_type, comment 
            FROM glucose 
            WHERE user_id = $1 AND date = $2
            ORDER BY time ASC
        ''', user_id, today)
    
    if not rows:
        return f"📊 Отчет по глюкозе за {today.strftime('%d-%m-%Y')}\n\nНет данных."
    
    report = f"📊 Отчет по глюкозе за {today.strftime('%d-%m-%Y')}\n\n"
    for row in rows:
        period_emoji = {"Утро": "🌅", "День": "☀️", "Вечер": "🌙"}
        emoji = period_emoji.get(row['period'], "")
        
        report += f"{emoji} {row['period']} {row['time'].strftime('%H:%M:%S')}: глюкоза {float(row['glucose_value'])}"
        if row['glucose_type']:
            report += f" ({row['glucose_type']})"
        if row['comment']:
            report += f"\n   📝 {row['comment']}"
        report += "\n\n"
    
    return report

# ==================== ГЛЮКОЗА - ОПРЕДЕЛЕНИЕ ТИПА ЗАМЕРА ====================
def detect_glucose_type(text):
    text_lower = text.lower()
    if "натощак" in text_lower or "на тощак" in text_lower:
        return "натощак"
    elif "через 2 часа" in text_lower or "после еды" in text_lower:
        return "через 2 часа после еды"
    elif "перед едой" in text_lower:
        return "перед едой"
    elif "перед сном" in text_lower:
        return "перед сном"
    elif "ночью" in text_lower or "ночь" in text_lower:
        return "ночью"
    else:
        now = datetime.now(MSK_PLUS_1)
        if 6 <= now.hour < 12:
            return "натощак"
        return "без указания"

# ==================== АДМИН КОМАНДЫ ====================
async def admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("Доступ запрещён.")
        return
    
    await update.message.reply_text(
        "👑 Админ панель\n\n"
        "Доступные команды:\n"
        "/users - список пользователей\n"
        "/users_excel - выгрузить пользователей в Excel\n"
        "/grant username дни - выдать доступ (5,7,30)\n"
        "/backup - создать резервную копию (SQL дамп)\n"
        "/restore - восстановить данные из SQL дампа\n"
        "/status - статус бота\n"
        "/test_remind - тестовая отправка напоминаний"
    )

async def admin_users(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("Доступ запрещён.")
        return
    
    users = await get_all_users()
    if not users:
        await update.message.reply_text("Нет пользователей.")
        return
    
    text = "👥 Список пользователей:\n\n"
    for uid, data in users.items():
        text += f"🆔 ID: {uid}\n"
        text += f"📝 Username: {data.get('username', '-')}\n"
        text += f"📅 Подключен: {data.get('joined', '-')}\n"
        text += f"🔒 Статус: {data.get('status', '-')}\n"
        text += f"📊 Дней: {data.get('days_count', '-')}\n"
        if data.get('access_until'):
            text += f"⏰ Доступ до: {data['access_until']}\n"
        if data.get('payment_info'):
            text += f"💳 Оплата: {data['payment_info']}\n"
        text += "-" * 30 + "\n"
    
    await update.message.reply_text(text)

async def admin_users_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("Доступ запрещён.")
        return
    
    users = await get_all_users()
    if not users:
        await update.message.reply_text("Нет пользователей.")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    
    headers = ["ID", "Username", "Дата подключения", "Статус", "Дней", "Доступ до", "Оплата", "Дата оплаты"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
    
    row = 2
    for uid, data in users.items():
        ws.cell(row=row, column=1, value=uid)
        ws.cell(row=row, column=2, value=data.get('username', '-'))
        ws.cell(row=row, column=3, value=data.get('joined', '-'))
        ws.cell(row=row, column=4, value=data.get('status', '-'))
        ws.cell(row=row, column=5, value=data.get('days_count', '-'))
        ws.cell(row=row, column=6, value=data.get('access_until', '-'))
        ws.cell(row=row, column=7, value=data.get('payment_info', '-'))
        ws.cell(row=row, column=8, value=data.get('payment_date', '-'))
        row += 1
    
    filename = f"users_{datetime.now(MSK_PLUS_1).strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    
    with open(filename, 'rb') as f:
        await update.message.reply_document(f, filename=filename)
    
    os.remove(filename)

async def admin_grant(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("Доступ запрещён.")
        return
    
    args = context.args
    if len(args) < 2:
        await update.message.reply_text(
            "📝 Используйте: /grant username дни\n"
            "Дни: 5, 7, 30\n"
            "Пример: /grant john 7"
        )
        return
    
    username = args[0]
    days = int(args[1])
    
    if days not in [5, 7, 30]:
        await update.message.reply_text("❌ Доступны дни: 5, 7, 30")
        return
    
    users = await get_all_users()
    found = None
    for uid, data in users.items():
        if data.get('username', '').lower() == username.lower():
            found = int(uid)
            break
    
    if found:
        await grant_access(found, days)
        await update.message.reply_text(f"✅ Пользователю {username} выдан доступ на {days} дней.")
    else:
        await update.message.reply_text(f"❌ Пользователь {username} не найден.")

async def backup_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Создаёт резервную копию БД (SQL дамп)"""
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("Доступ запрещён.")
        return
    
    try:
        timestamp = datetime.now(MSK_PLUS_1).strftime("%Y%m%d_%H%M%S")
        filename = f"backup_{timestamp}.sql"
        
        # Делаем дамп с помощью pg_dump
        import subprocess
        result = subprocess.run(
            ['pg_dump', DATABASE_URL, '--clean', '--if-exists', '-f', filename],
            capture_output=True,
            text=True
        )
        
        if result.returncode != 0:
            await update.message.reply_text(f"❌ Ошибка при создании дампа: {result.stderr}")
            return
        
        with open(filename, 'rb') as f:
            await update.message.reply_document(
                document=f,
                filename=filename,
                caption=f"📦 Резервная копия БД от {datetime.now(MSK_PLUS_1).strftime('%d-%m-%Y %H:%M:%S')}"
            )
        
        os.remove(filename)
        print(f"Создана резервная копия БД: {filename}")
        
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка при создании резервной копии: {e}")

async def restore_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Восстанавливает данные из SQL дампа"""
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("Доступ запрещён.")
        return
    
    await update.message.reply_text(
        "📤 Отправьте SQL дамп (созданный командой /backup).\n\n"
        "⚠️ ВНИМАНИЕ: текущие данные будут ПЕРЕЗАПИСАНЫ!"
    )
    context.user_data['awaiting_restore'] = 'sql'

async def status_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Проверка статуса бота (только админ)"""
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("Доступ запрещён.")
        return
    
    status = "работает" if context.application.job_queue else "НЕ РАБОТАЕТ"
    await update.message.reply_text(f"🤖 Статус бота:\n\nJobQueue: {status}\nБаза данных: ✅ подключена")

async def test_remind_all(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Тестовая отправка напоминания всем (только админ)"""
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("Доступ запрещён.")
        return
    
    await update.message.reply_text("🔄 Отправляю тестовые напоминания всем пользователям...")
    
    users = await get_all_users()
    sent = 0
    for uid, data in users.items():
        if await check_access(int(uid)):
            try:
                await context.bot.send_message(
                    chat_id=int(uid),
                    text="🧪 ТЕСТ: Напоминание работает! Если вы это видите — бот исправен."
                )
                sent += 1
            except:
                pass
    
    await update.message.reply_text(f"✅ Тестовое напоминание отправлено {sent} пользователям")

# ==================== ОСНОВНЫЕ КОМАНДЫ ====================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    username = update.effective_user.username or update.effective_user.first_name
    
    is_new = await add_user(user_id, username)
    
    if is_new and not is_admin(user_id):
        now = datetime.now(MSK_PLUS_1)
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"🆕 Новый пользователь!\n\n"
                 f"👤 Username: @{username}\n"
                 f"🆔 ID: {user_id}\n"
                 f"📅 Дата: {now.strftime('%d-%m-%Y %H:%M:%S')}"
        )
    
    await check_and_send_3day_reminder(user_id, context.application)
    
    await update.message.reply_text(
        "📊 Я помогу вести журнал вашего артериального давления и уровня глюкозы.\n\n"
        "📝 Форматы ввода давления:\n"
        "• 120 80 - давление\n"
        "• 120 80 68 - давление и пульс\n"
        "• 120 80 выпил таблетку - с комментарием\n\n"
        "📝 Форматы ввода глюкозы:\n"
        "• 5.5 - глюкоза (период определится автоматически)\n"
        "• 5.5 натощак - глюкоза с типом замера\n"
        "• 5.5 через 2 часа после еды - глюкоза с типом замера\n"
        "• 9,2 инсулин 10 - глюкоза и доза инсулина\n\n"
        "🌅 Бот сам определит время суток (Утро, День, Вечер)\n"
        "💾 Все данные хранятся в защищённой базе данных\n\n"
        "Команды:\n"
        "/table - получить Excel файл (2 листа: давление и глюкоза)\n"
        "/report - отчет по давлению за сегодня\n"
        "/glucose_report - отчет по глюкозе за сегодня\n"
        "/help - помощь"
    )

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (
        "📖 Помощь\n\n"
        "Как пользоваться:\n"
        "1. Отправьте показания давления или глюкозы\n"
        "2. Бот сам определит время суток (Утро, День, Вечер)\n"
        "3. Каждый замер сохраняется отдельной строкой\n\n"
        "Форматы ввода давления:\n"
        "• 130 85 - давление\n"
        "• 130 85 72 - давление и пульс\n"
        "• 130 85 выпил таблетку - с комментарием\n\n"
        "Форматы ввода глюкозы:\n"
        "• 5.5 - глюкоза\n"
        "• 5.5 натощак - глюкоза с типом замера\n"
        "• 5.5 через 2 часа после еды\n"
        "• 9,2 инсулин 10 - глюкоза и доза инсулина\n\n"
        "Рекомендации по измерению глюкозы:\n"
        "• Утром натощак\n"
        "• Перед каждым приёмом пищи\n"
        "• Через 2 часа после еды\n"
        "• Перед сном\n\n"
        "Целевые показатели:\n"
        "• Натощак: 4.0–7.0 ммоль/л\n"
        "• Через 2 часа после еды: менее 10.0 ммоль/л\n\n"
        "Команды:\n"
        "/table - Excel файл (давление и глюкоза)\n"
        "/report - отчет по давлению за сегодня\n"
        "/glucose_report - отчет по глюкозе за сегодня\n\n"
        f"📢 <a href='https://t.me/+MAuGbcnBQmgxZTIy'>Больше наших ботов в канале</a>"
    )
    
    await update.message.reply_text(text, parse_mode="HTML", disable_web_page_preview=True)

async def report_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    report = await get_today_pressure_report(user_id)
    await update.message.reply_text(report)

async def glucose_report_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    report = await get_today_glucose_report(user_id)
    await update.message.reply_text(report)

async def table_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    
    await update.message.reply_text("🔄 Генерирую ваш Excel-файл...")
    
    try:
        filename = await generate_pressure_excel(user_id)
        
        with open(filename, 'rb') as f:
            await update.message.reply_document(
                document=f,
                filename="medical_journal.xlsx",
                caption="📊 Ваш медицинский журнал (давление и глюкоза)\n\n"
                        "Все ваши данные — только ваши. Другие пользователи не видят их."
            )
        
        os.remove(filename)
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка при создании файла: {e}")

async def handle_pressure_glucose(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    username = update.effective_user.username or update.effective_user.first_name
    
    await add_user(user_id, username)
    
    if not await check_access(user_id):
        await update.message.reply_text(
            f"⛔ Доступ временно приостановлен.\nСвяжитесь с администратором @{ADMIN_USERNAME}"
        )
        return
    
    await check_and_send_3day_reminder(user_id, context.application)
    await update_user_days(user_id)
    
    text = update.message.text.strip()
    text_lower = text.lower()
    
    # Проверяем, есть ли в тексте упоминание инсулина
    is_insulin = any(word in text_lower for word in ["инсулин", "инсулина", "ед", "единиц", "iu", "ме", "мед", "единица"])
    
    # Ищем все числа (включая десятичные)
    numbers = re.findall(r'\d+[.,]?\d*', text)
    numbers = [float(n.replace(',', '.')) for n in numbers]
    
    # Если есть инсулин или одно число 1-30 — это глюкоза
    if is_insulin or (len(numbers) >= 1 and 1 <= numbers[0] <= 30):
        # Это глюкоза
        glucose = None
        insulin = None
        comment_parts = []
        
        for n in numbers:
            if glucose is None and 1 <= n <= 30:
                glucose = n
            elif insulin is None and n > 0:
                insulin = int(n) if n.is_integer() else n
        
        # Если не нашли глюкозу, берём первое число
        if glucose is None and numbers:
            glucose = numbers[0]
        
        glucose_type = detect_glucose_type(text)
        period = get_period_by_time()
        
        # Удаляем числа и ключевые слова из комментария
        comment = re.sub(r'\d+[.,]?\d*', '', text)
        comment = re.sub(r'инсулин|инсулина|ед|единиц|iu|ме|мед', '', comment, flags=re.IGNORECASE)
        comment = re.sub(r'натощак|через 2 часа после еды|перед едой|перед сном|ночью', '', comment, flags=re.IGNORECASE)
        comment = re.sub(r'[\s/]+', ' ', comment).strip()
        
        # Если есть доза инсулина, добавляем в комментарий
        if insulin:
            if comment:
                comment = f"{comment}, инсулин {insulin} ед."
            else:
                comment = f"инсулин {insulin} ед."
        
        await save_glucose_to_db(user_id, period, glucose, glucose_type, comment)
        
        period_emoji = {"Утро": "🌅", "День": "☀️", "Вечер": "🌙"}
        now = datetime.now(MSK_PLUS_1)
        
        response = f"✅ Записано! {period_emoji.get(period, '')} {period}: глюкоза {glucose}"
        if glucose_type != "без указания":
            response += f" ({glucose_type})"
        if insulin:
            response += f", инсулин {insulin} ед."
        if comment and not comment.startswith("инсулин"):
            response += f"\n📝 {comment}"
        response += f"\n📅 {now.strftime('%d-%m-%Y %H:%M:%S')}"
        
        await update.message.reply_text(response)
        return
    
    # Давление
    systolic = None
    diastolic = None
    pulse = None
    
    slash_match = re.search(r'(\d{2,3})/(\d{2,3})', text)
    if slash_match:
        systolic = int(slash_match.group(1))
        diastolic = int(slash_match.group(2))
        numbers = [n for n in numbers if n not in [systolic, diastolic]]
    elif len(numbers) >= 2:
        systolic = int(numbers[0])
        diastolic = int(numbers[1])
        numbers = numbers[2:]
    
    if not systolic or not diastolic:
        await update.message.reply_text(
            "❌ Не понял. Примеры:\n"
            "120 80 - давление\n"
            "120 80 68 - давление и пульс\n"
            "5.5 - глюкоза\n"
            "9,2 инсулин 10 - глюкоза и инсулин\n"
            "120 80 выпил таблетку - с комментарием"
        )
        return
    
    for n in numbers:
        if 40 <= n <= 150:
            pulse = int(n)
            break
    
    comment = re.sub(r'\d+[.,]?\d*', '', text)
    comment = re.sub(r'[\s/]+', ' ', comment).strip()
    
    period = get_period_by_time()
    await save_pressure_to_db(user_id, period, systolic, diastolic, pulse, comment)
    
    period_emoji = {"Утро": "🌅", "День": "☀️", "Вечер": "🌙"}
    now = datetime.now(MSK_PLUS_1)
    
    response = f"✅ Записано! {period_emoji.get(period, '')} {period}: {systolic}/{diastolic}"
    if pulse:
        response += f", пульс {pulse}"
    if comment:
        response += f"\n📝 {comment}"
    response += f"\n📅 {now.strftime('%d-%m-%Y %H:%M:%S')}"
    
    await update.message.reply_text(response)

# ==================== ВОССТАНОВЛЕНИЕ ИЗ SQL ДАМПА ====================
async def handle_restore_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик загруженного SQL файла для восстановления"""
    if not is_admin(update.effective_user.id):
        return
    
    if context.user_data.get('awaiting_restore') != 'sql':
        return
    
    document = update.message.document
    if not document or not document.file_name.endswith('.sql'):
        await update.message.reply_text("❌ Пожалуйста, отправьте SQL дамп (созданный командой /backup)")
        return
    
    try:
        file = await context.bot.get_file(document.file_id)
        
        temp_file = f"temp_restore_{datetime.now(MSK_PLUS_1).strftime('%Y%m%d_%H%M%S')}.sql"
        await file.download_to_drive(temp_file)
        
        import subprocess
        result = subprocess.run(
            ['psql', DATABASE_URL, '-f', temp_file],
            capture_output=True,
            text=True
        )
        
        os.remove(temp_file)
        
        if result.returncode != 0:
            await update.message.reply_text(f"❌ Ошибка при восстановлении: {result.stderr}")
        else:
            await update.message.reply_text(f"✅ Данные восстановлены из файла {document.file_name}")
        
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка при восстановлении: {e}")
    finally:
        context.user_data['awaiting_restore'] = None

# ==================== НАПОМИНАНИЯ ====================
async def send_scheduled_reminder(context: ContextTypes.DEFAULT_TYPE):
    """Отправляет напоминания всем активным пользователям"""
    now_time = datetime.now(MSK_PLUS_1)
    current_hour = now_time.hour
    
    if current_hour not in [8, 14, 20]:
        return
    
    print(f"[{now_time.strftime('%Y-%m-%d %H:%M:%S')}] Запуск напоминаний, час: {current_hour}")
    
    users = await get_all_users()
    sent_count = 0
    active_count = 0
    
    for uid, data in users.items():
        if await check_access(int(uid)):
            active_count += 1
            try:
                await context.bot.send_message(
                    chat_id=int(uid),
                    text="🔔 Напоминание: пора измерить давление и глюкозу!\n\n"
                         "Просто отправьте мне показания:\n"
                         "• 120 80 - давление\n"
                         "• 120 80 68 - давление и пульс\n"
                         "• 5.5 - глюкоза\n"
                         "• 5.5 натощак - глюкоза с типом замера\n"
                         "• 9,2 инсулин 10 - глюкоза и доза инсулина"
                )
                sent_count += 1
                print(f"  → Напоминание отправлено пользователю {uid}")
            except Exception as e:
                print(f"  ✗ Ошибка отправки пользователю {uid}: {e}")
    
    print(f"Активных пользователей: {active_count}, отправлено напоминаний: {sent_count}")

# ==================== КОМАНДЫ МЕНЮ ====================
async def set_commands(app):
    admin_commands = [
        BotCommand("start", "Главное меню"),
        BotCommand("table", "Excel журнал (давление+глюкоза)"),
        BotCommand("report", "Отчет по давлению за сегодня"),
        BotCommand("glucose_report", "Отчет по глюкозе за сегодня"),
        BotCommand("help", "Помощь"),
        BotCommand("admin", "Админ панель"),
        BotCommand("users", "Список пользователей"),
        BotCommand("users_excel", "Выгрузить пользователей в Excel"),
        BotCommand("grant", "Выдать доступ (username дни)"),
        BotCommand("backup", "Резервная копия БД"),
        BotCommand("restore", "Восстановить БД"),
        BotCommand("status", "Статус бота"),
        BotCommand("test_remind", "Тест напоминаний"),
    ]
    
    default_commands = [
        BotCommand("start", "Главное меню"),
        BotCommand("table", "Excel журнал (давление+глюкоза)"),
        BotCommand("report", "Отчет по давлению за сегодня"),
        BotCommand("glucose_report", "Отчет по глюкозе за сегодня"),
        BotCommand("help", "Помощь"),
    ]
    
    await app.bot.set_my_commands(default_commands)
    await app.bot.set_my_commands(admin_commands, scope=BotCommandScopeChat(chat_id=ADMIN_ID))

# ==================== ЗАПУСК ====================
async def main():
    """Основная асинхронная функция запуска бота"""
    global db_pool
    
    # Инициализация БД
    await init_db()
    
    # Создаём приложение
    app = Application.builder().token(TOKEN).build()
    
    if app.job_queue is None:
        print("❌ ОШИБКА: JobQueue не создан! Напоминания работать не будут")
    else:
        print("✅ JobQueue создан успешно")
        print(f"   Текущее время сервера: {datetime.now(MSK_PLUS_1).strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"   Часовой пояс: Europe/Samara (МСК+1)")
    
    # Регистрируем обработчики
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("report", report_command))
    app.add_handler(CommandHandler("glucose_report", glucose_report_command))
    app.add_handler(CommandHandler("table", table_command))
    
    app.add_handler(CommandHandler("admin", admin_panel))
    app.add_handler(CommandHandler("users", admin_users))
    app.add_handler(CommandHandler("users_excel", admin_users_excel))
    app.add_handler(CommandHandler("grant", admin_grant))
    app.add_handler(CommandHandler("backup", backup_command))
    app.add_handler(CommandHandler("restore", restore_command))
    app.add_handler(CommandHandler("status", status_command))
    app.add_handler(CommandHandler("test_remind", test_remind_all))
    
    app.add_handler(MessageHandler(filters.Document.ALL, handle_restore_file))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_pressure_glucose))
    
    # Устанавливаем команды меню
    await set_commands(app)
    
    # Настройка напоминаний
    job_queue = app.job_queue
    if job_queue:
        job_queue.run_daily(send_scheduled_reminder, time(5, 0))
        job_queue.run_daily(send_scheduled_reminder, time(11, 0))
        job_queue.run_daily(send_scheduled_reminder, time(17, 0))
        print("Напоминания: 8:00, 14:00, 20:00 (МСК+1)")
    else:
        print("ОШИБКА: job_queue не создан! Напоминания работать не будут")
    
    print("🤖 Бот запущен")
    
    # Запускаем бота
    await app.initialize()
    await app.start()
    await app.run_polling()

if __name__ == "__main__":
    asyncio.run(main())